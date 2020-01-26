import smtplib
import openpyxl
from string import Template
import getpass 

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


#Function to retrive names and email addresses to a list and return them
def excel_to_list(filename):
    names=[]
    emails=[] 
    wb_obj = openpyxl.load_workbook(filename) 
    sheet_obj = wb_obj.active 
    max_row = sheet_obj.max_row
    for i in range(2,max_row+1):
        name_obj = sheet_obj.cell(row = i, column = 1)
        names.append(name_obj.value)
        email_obj = sheet_obj.cell(row = i, column = 2)
        emails.append(email_obj.value)
    return names,emails
#Function to load the template file
def template_file(filename):
    with open(filename, 'r', encoding='utf-8') as template_file:
        template_file_content = template_file.read()
    return Template(template_file_content)

def main():
    names,emails=excel_to_list("participents.xlsx")
    template_content=template_file("template.txt")
    s = smtplib.SMTP(host='smtp.gmail.com', port=587)
    s.starttls()
    MY_ADDRESS=input("ENter your Email:")
    MY_PASSWORD=getpass.getpass()
    
    s.login(MY_ADDRESS,MY_PASSWORD)
    for names,emails in zip(names,emails):
        msg=MIMEMultipart()
        message=template_content.substitute(PERSON_NAME=names.title())

        msg['From']=MY_ADDRESS
        msg['To']=emails
        msg['Subject']="Hi"+str(names)

        msg.attach(MIMEText(message, 'plain'))
        certificateFile=str(names)+".pdf"
        filename = "Certificate.pdf"
        attachment = open(certificateFile, "rb") 
        p = MIMEBase('application', 'octet-stream')
        p.set_payload((attachment).read()) 
        encoders.encode_base64(p)
        p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
        msg.attach(p)

        s.send_message(msg)
        print("Message sent to ",names," having email",emails)
        

        del msg
    s.quit()

if __name__=='__main__':
    main()
